import pprint
from openpyxl import load_workbook

s1 = "baza_cvs/1. БС_модули_блоки.csv"
s2 = "baza_cvs/2. Группы_бо_во_фб.csv"
s3 = "baza_cvs/4.Ожидаемые поставки_Декабрь2020.csv"


def get_text1(s) -> dict:
    """ Достает из cvs данные и формирует из них список
    :param s: путь к файлу cvs из которого получаем данные
    :return: словарь, где ключ это номер класса оборудования, а значение это спсок оборудования
    """
    with open(s, 'r', encoding='utf-8') as file:
        stroka = file.readlines()
        dct = {}
        for i in stroka[1:]:
            s = i.split(';')
            s = [i.strip() for i in s]
            s_item = s[0].strip().split('.')[0]
            if s_item not in dct:
                dct[s_item] = [s[1]]
            elif s_item in dct:
                dct[s_item] = dct[s_item] + [s[1]]
    return dct


def get_text2(s):
    '''Достает из cvs данные и формирует из них список
    :param s: путь к файлу cvs из которого получаем данные
    :return: словарь, где ключ это номер класса оборудования, а значение это категория оборудования
            ВО демонтаж, не БС демонтаж,не БС, ВО, БС, ФБ демонтаж
    '''
    with open(s, 'r', encoding='utf-8') as file:
        stroka = file.readlines()
        dct = {}
        for i in stroka[1:]:
            s = i.split(';')
            s = [i.strip() for i in s]
            s_item = s[0].strip().split('.')[0]
            if s_item not in dct:
                dct[int(s_item)] = [s[2]]
            elif s_item in dct:
                dct[int(s_item)] += [s[2]]
    return dct


def get_text3(s):
    with open(s, 'r', encoding='utf-8') as file:
        stroka = file.readlines()
        dct_postavka = {}
        dct_sclad = {}
        for i in stroka[1:]:
            s = i.split(';')
            s = [i.strip() for i in s]
            sklad = s[0]
            s_kod = s[1].strip().split('.')[0]
            s_v = s[2]
            data = s[3]
            if s_kod not in dct_postavka:
                dct_postavka[s_kod] = [data]
            elif s_kod in dct_postavka:
                dct_postavka[s_kod] = dct_postavka[s_kod] + [data]
    return dct_postavka


def get_united(dct1: dict, dct2: dict):
    dct_united = {}
    for i, j in dct1.items():
        for k, v in dct2.items():
            if i == k:
                dct_united[''.join(v)] = j
    return dct_united


dct1 = get_text1(s1)
dct2 = get_text2(s2)
dct3 = get_text3(s3)

dict_united = get_united(dct1, dct2)
# print(dct1)
# print(dct2)
pprint.pprint(dct3)

# wb = load_workbook('result.xlsx')
# sheet = wb.get_sheet_by_name('Лист1')
#
# sheet['A1'] = '1'
# wb.save('test4.xlsx')


def create_file(name_file) -> None:
    '''Создает файл xlsx с именем name_file'''
    pass



def write_data_to_a_file(s: dict, name_file: str) -> None:
    '''Из словаря взять данные и записать в файл эксель. Ключ в колонке А, а значения в В
    '''
    wb = load_workbook(name_file)
    sheet = wb['Лист1']
    sheet['A1'] = '1'

    count = 1
    for key, value in s.items():
        sheet['A' + str(count)] = int(key)
        for j in value:
            sheet['B' + str(count)] = j
            count += 1
    wb.save(name_file)

# file = 'test.xlsx'
# write_data_to_a_file(dct1, file)
